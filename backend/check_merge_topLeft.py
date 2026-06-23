"""Verifica cuáles celdas son top-left de sus rangos combinados."""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path

TEMPLATE = Path(__file__).parent / "assets" / "solicitud_template.xlsx"

CELLS = [
    "I10", "F14", "T14", "F16", "T16",
    "G19", "G21", "G26", "Q26",
    "G33", "S33",
    "C37", "G37", "S37",
    "G39",
    "G49", "S49", "G52",
    "B70", "G70",
    "C83", "G83", "O83", "T83",
]

wb = openpyxl.load_workbook(str(TEMPLATE))
ws = wb["Solicitud de pago"]

print("=== Revisando celdas vs merged ranges ===\n")
print(f"{'Celda':8} {'Tipo':14} {'Rango merge':16} {'Top-left':10} {'Correcto?'}")
print("-" * 65)

for cell_id in CELLS:
    cell = ws[cell_id]
    cell_type = type(cell).__name__

    # Find if this cell is in any merged range
    merge_range = None
    top_left = None
    for rng in ws.merged_cells.ranges:
        # Check if cell_id falls within this range
        col_letter = ''.join(c for c in cell_id if c.isalpha())
        row_num = int(''.join(c for c in cell_id if c.isdigit()))
        col_num = column_index_from_string(col_letter)

        if (rng.min_row <= row_num <= rng.max_row and
                rng.min_col <= col_num <= rng.max_col):
            merge_range = str(rng)
            tl_col = get_column_letter(rng.min_col)
            top_left = f"{tl_col}{rng.min_row}"
            break

    is_correct = (top_left == cell_id) if top_left else (cell_type == "Cell")
    marker = "OK" if is_correct else "FIX -> usar " + (top_left or "?")

    print(f"{cell_id:8} {cell_type:14} {merge_range or '(no merge)':16} "
          f"{top_left or cell_id:10} {marker}")
