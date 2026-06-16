"""
Importador de PagosIT.xlsx -> Supabase Postgres.

Lee la hoja con servicios recurrentes y sus columnas mensuales (abril/mayo/junio)
y carga proveedores, empresas/CC, servicios y un pago por mes:
  celda True  -> PAGADO
  celda False -> PENDIENTE
  celda vacía -> se omite (no aplica ese mes)

Reinicia (TRUNCATE) las tablas de datos antes de cargar, así es re-ejecutable.
Los bancos se cargan de la lista de referencia (no vienen en el Excel).

Uso:
    cd backend
    .venv/Scripts/python import_excel.py ["ruta\\PagosIT.xlsx"]
"""
from __future__ import annotations
import re
import sys
from datetime import date

import openpyxl
import psycopg

from app.config import get_settings
from seed import BANCOS, EMP_MAP, SUC_MAP, norm_emp, norm_suc

DEFAULT_XLSX = r"c:\Users\Analista de sistemas\OneDrive - SELECT SHOP MB SA DE CV\New era\PagosIT.xlsx"

# Columnas (0-indexadas) de la hoja
C_PROV, C_CUENTA, C_CONCEPTO, C_MONTO = 0, 1, 2, 3
C_EMPRESA, C_SUCURSAL, C_GASTO, C_AREA, C_DIA = 4, 5, 6, 7, 8
C_ABR, C_MAY, C_JUN, C_COMENT = 9, 10, 11, 12

MESES_COLS = [(C_ABR, 4, "Abril"), (C_MAY, 5, "Mayo"), (C_JUN, 6, "Junio")]


def limpiar_txt(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def parse_monto(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(",", "")          # coma = miles, punto = decimal
    elif "," in s:
        s = s.replace(",", ".")         # coma = decimal
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def parse_dia(v):
    m = re.search(r"\d+", str(v or ""))
    return int(m.group()) if m else None


def importar(conn, xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    cur = conn.cursor()

    # ── Reset ────────────────────────────────────────────
    cur.execute("TRUNCATE pagos, servicios_recurrentes, empresas_cc, proveedores, bancos "
                "RESTART IDENTITY CASCADE")

    # ── Bancos de referencia ─────────────────────────────
    for nom, pref in BANCOS:
        cur.execute("INSERT INTO bancos (nombre,prefijo_clabe) VALUES (%s,%s) "
                    "ON CONFLICT (nombre) DO NOTHING", (nom, pref))

    prov_cache: dict[str, int] = {}
    emp_cache: dict[tuple, int] = {}

    def get_prov(nombre: str) -> int:
        if nombre not in prov_cache:
            row = cur.execute(
                "INSERT INTO proveedores (nombre,banco,clabe) VALUES (%s,'BBVA','') RETURNING id",
                (nombre,)).fetchone()
            prov_cache[nombre] = row[0]
        return prov_cache[nombre]

    def get_emp(empresa, sucursal, cc, direccion, gasto) -> int:
        key = (empresa, sucursal, cc)
        if key not in emp_cache:
            row = cur.execute(
                "INSERT INTO empresas_cc (empresa,sucursal,centro_costos,direccion,tipo_gasto) "
                "VALUES (%s,%s,%s,%s,%s) RETURNING id",
                (empresa, sucursal, cc, direccion, gasto or "GASTO")).fetchone()
            emp_cache[key] = row[0]
        return emp_cache[key]

    n_srv = n_pag = n_skip = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        prov_raw = limpiar_txt(r[C_PROV]) if len(r) > C_PROV else ""
        if not prov_raw:
            n_skip += 1
            continue

        cuenta = limpiar_txt(r[C_CUENTA]) if len(r) > C_CUENTA else ""
        concepto = limpiar_txt(r[C_CONCEPTO]) if len(r) > C_CONCEPTO else ""
        monto = parse_monto(r[C_MONTO]) if len(r) > C_MONTO else 0.0
        empresa = norm_emp(limpiar_txt(r[C_EMPRESA])) if len(r) > C_EMPRESA else ""
        sucursal = norm_suc(limpiar_txt(r[C_SUCURSAL])) if len(r) > C_SUCURSAL else ""
        gasto = limpiar_txt(r[C_GASTO]) if len(r) > C_GASTO else ""
        area = limpiar_txt(r[C_AREA]) if len(r) > C_AREA else ""
        dia = parse_dia(r[C_DIA]) if len(r) > C_DIA else None

        prov_id = get_prov(prov_raw)
        emp_id = get_emp(empresa, sucursal, area, area, gasto)

        desc = concepto or cuenta or prov_raw[:25]
        srv_id = cur.execute(
            """INSERT INTO servicios_recurrentes
               (proveedor_id,empresa_cc_id,descripcion,no_cuenta_servicio,
                tipo,dia_limite,monto_base,iva,activo)
               VALUES (%s,%s,%s,%s,%s,%s,%s,0,true) RETURNING id""",
            (prov_id, emp_id, desc, cuenta, concepto, dia, monto)).fetchone()[0]
        n_srv += 1

        base_motivo = cuenta or concepto or prov_raw
        for col, mes_num, mes_nom in MESES_COLS:
            cell = r[col] if len(r) > col else None
            if cell is None or cell == "":
                continue
            estatus = "PAGADO" if cell is True else "PENDIENTE"
            dia_real = min(dia, 28) if dia else 15
            fecha_proc = f"2026-{mes_num:02d}-{dia_real:02d}"
            motivo = f"SERV {base_motivo} {mes_nom.upper()} 2026"
            cur.execute(
                """INSERT INTO pagos
                   (servicio_id,proveedor_nombre,empresa,sucursal,centro_costos,direccion,
                    motivo_pago,monto_total,banco,no_cuenta,
                    mes_presupuesto,mes_pago,mes,anio,estatus,fecha_proceso)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'BBVA',%s,%s,%s,%s,%s,%s,%s)""",
                (srv_id, prov_raw, empresa, sucursal, area, area,
                 motivo, monto, cuenta, mes_nom, mes_nom, mes_num, 2026, estatus, fecha_proc))
            n_pag += 1

    conn.commit()
    print(f"Importación OK desde: {xlsx_path}")
    print(f"  proveedores: {len(prov_cache)}")
    print(f"  empresas/CC: {len(emp_cache)}")
    print(f"  servicios:   {n_srv}")
    print(f"  pagos:       {n_pag}")
    print(f"  filas omitidas (sin proveedor): {n_skip}")


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    dsn = get_settings().database_url
    if not dsn:
        raise SystemExit("DATABASE_URL no configurada.")
    with psycopg.connect(dsn, connect_timeout=20) as conn:
        importar(conn, xlsx)


if __name__ == "__main__":
    main()
