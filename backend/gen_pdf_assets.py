"""
gen_pdf_assets.py — genera los assets de PDF para producción.

Corre UNA SOLA VEZ en Windows con Excel instalado.

Genera:
  backend/assets/solicitud_bg.pdf   — formulario en blanco (fondo exacto del Excel real)
  backend/assets/cell_coords.json   — coordenadas de cada celda de datos en el PDF
"""
import os, json, shutil, tempfile
from pathlib import Path

import win32com.client as w32
import openpyxl
import fitz  # PyMuPDF

BASE = Path(__file__).parent
TEMPLATE_XLSX = BASE / "assets" / "solicitud_template.xlsx"
BG_PDF_OUT   = BASE / "assets" / "solicitud_bg.pdf"
COORDS_OUT   = BASE / "assets" / "cell_coords.json"

# Celdas de datos y su campo semántico
CELLS = [
    "I10", "F14", "T14", "F16", "T16",
    "G19", "G21", "G26", "P26",
    "G33", "S33",
    "C37", "G37", "S37",
    "G39",
    "G49", "S49", "G52",
    "B70", "G70",
    "C83", "G83", "O83", "T83",
]

# Marcadores únicos: "z001", "z002", … — no aparecen en etiquetas del formulario
MARKERS = {cell: f"z{i:03d}" for i, cell in enumerate(CELLS, 1)}

tmp_dir = Path(tempfile.mkdtemp())
print(f"Directorio temporal: {tmp_dir}\n")

# ── 1. PDF de fondo (formulario vacío, exportado por Excel) ──────────
print("=== 1/3 Generando PDF de fondo ===")
xl = w32.Dispatch("Excel.Application")
xl.DisplayAlerts = False
xl.Visible = False

wbb = xl.Workbooks.Open(str(TEMPLATE_XLSX.resolve()))
bg_tmp = str(tmp_dir / "bg.pdf")
wbb.Sheets(1).ExportAsFixedFormat(0, bg_tmp)   # 0 = xlTypePDF
wbb.Close(False)
print(f"  Exportado: {bg_tmp} ({os.path.getsize(bg_tmp)} bytes)")

# ── 2. PDF de calibración (con marcadores en cada celda de datos) ────
print("\n=== 2/3 Generando PDF de calibración ===")
wb = openpyxl.load_workbook(str(TEMPLATE_XLSX))
ws = wb["Solicitud de pago"]

font_sizes = {}
for cell_id, marker in MARKERS.items():
    cell = ws[cell_id]
    # Guardar tamaño de fuente
    sz = None
    if cell.font and cell.font.size:
        sz = float(cell.font.size)
    font_sizes[cell_id] = sz or 8.0
    # Escribir marcador
    ws[cell_id] = marker

cal_xlsx_tmp = str(tmp_dir / "cal.xlsx")
wb.save(cal_xlsx_tmp)

wbc = xl.Workbooks.Open(os.path.abspath(cal_xlsx_tmp))
cal_tmp = str(tmp_dir / "cal.pdf")
wbc.Sheets(1).ExportAsFixedFormat(0, cal_tmp)
wbc.Close(False)
xl.Quit()
print(f"  Exportado: {cal_tmp} ({os.path.getsize(cal_tmp)} bytes)")

# ── 3. Calibrar posiciones con PyMuPDF ──────────────────────────────
print("\n=== 3/3 Calibrando posiciones ===")
doc = fitz.open(cal_tmp)
page = doc[0]
print(f"  Tamaño de página: {page.rect}")

coords = {}
not_found = []

for cell_id, marker in MARKERS.items():
    hits = page.search_for(marker)
    if hits:
        r = hits[0]
        coords[cell_id] = {
            "x0": round(r.x0, 2),
            "y0": round(r.y0, 2),
            "x1": round(r.x1, 2),
            "y1": round(r.y1, 2),
            "sz": font_sizes.get(cell_id, 8.0),
        }
        print(f"  {cell_id:5s} ('{marker}'): x0={r.x0:6.1f} y0={r.y0:6.1f}  "
              f"w={r.width:5.1f} h={r.height:4.1f}  sz={font_sizes.get(cell_id,8)}")
    else:
        not_found.append(f"{cell_id} ('{marker}')")

if not_found:
    print(f"\n  ADVERTENCIA — no encontrados:")
    for nf in not_found:
        print(f"    {nf}")

doc.close()

# ── 4. Guardar outputs ───────────────────────────────────────────────
shutil.copy2(bg_tmp, BG_PDF_OUT)
COORDS_OUT.write_text(json.dumps(coords, indent=2, ensure_ascii=False))
shutil.rmtree(tmp_dir, ignore_errors=True)

print(f"\n=== DONE ===")
print(f"  solicitud_bg.pdf  : {BG_PDF_OUT.stat().st_size} bytes")
print(f"  cell_coords.json  : {len(coords)} celdas calibradas")
if not_found:
    print(f"  REVISAR: {len(not_found)} celdas no encontradas (ver arriba)")
