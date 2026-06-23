"""Prueba que openpyxl puede escribir en las celdas del template."""
import openpyxl
from pathlib import Path

tpl = Path(__file__).parent / "assets" / "solicitud_template.xlsx"
wb = openpyxl.load_workbook(str(tpl))
ws = wb["Solicitud de pago"]

ws["I10"] = "GOLDEN YEARS MANAGEMENT"
ws["F14"] = "TELEFONOS DE MEXICO"
ws["T14"] = "23/06/2026"
ws["F16"] = "CC-001"
ws["T16"] = "BLVD. MANUEL AVILA CAMACHO 2900"
ws["G19"] = "PROVEEDOR EJEMPLO SA DE CV"
ws["G21"] = "SERVICIO TELEFONICO JUN 2026"
ws["G26"] = "ABC-123456"
ws["G33"] = "DOS MIL PESOS 00/100 M.N."
ws["S33"] = 2000.00
ws["C37"] = "BANAMEX"
ws["G37"] = "002010077777777771"
ws["S37"] = "12345678"
ws["G39"] = "Sin observaciones"
ws["G49"] = "JUNIO"
ws["S49"] = "JUNIO"
ws["G52"] = "CC-001"
ws["B70"] = "LILLY ESTEFANY ARROYO HUERTA"
ws["G70"] = "BRUNO CASTANEDA ROVIRA"
ws["C83"] = "DENIS TOLENTINO"
ws["G83"] = "PRESUPUESTOS"
ws["O83"] = "VICTOR NUNEZ"
ws["T83"] = "STEVEN MARCOVICH"

out = tpl.parent / "solicitud_test_filled.xlsx"
wb.save(str(out))
print("OK - celdas escritas sin errores")
print(f"I10 = {ws['I10'].value}")
print(f"S33 = {ws['S33'].value}")
print(f"G37 = {ws['G37'].value}")
print(f"Guardado en: {out}")
