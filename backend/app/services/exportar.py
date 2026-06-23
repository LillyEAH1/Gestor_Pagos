"""
exportar.py (backend) — generación de PDF de Solicitud de Pago.

Usa como fondo el JPG generado desde formato.xlsx (fit-to-1-page).
Coordenadas x/y calibradas contra el PDF relleno del Excel exportado.
RL y = 792 - y_top  (PyMuPDF da y desde arriba, ReportLab desde abajo)
"""
from __future__ import annotations
import io
import json
from pathlib import Path
from datetime import date

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

PW, PH = letter          # 612 × 792 pts
NEGRO = colors.black

_ASSETS = Path(__file__).resolve().parents[2] / "assets" / "logos"


def _plantilla() -> str:
    p = _ASSETS / "formato2_page-0001.jpg"
    return str(p) if p.exists() else ""


def _get_logo(empresa: str):
    if not empresa:
        return None
    mf = _ASSETS / "logo_map.json"
    if not mf.exists():
        return None
    try:
        m = json.loads(mf.read_text(encoding="utf-8"))
        eu = empresa.upper()
        for k, v in m.items():
            if k.upper() in eu or eu[:12] in k.upper():
                p = _ASSETS / v
                if p.exists():
                    return str(p)
    except Exception:
        pass
    return None


def exportar_solicitud_pdf(datos: dict) -> bytes:
    """Genera el PDF en memoria y devuelve los bytes. NUNCA pone empresa por default."""

    def g(k: str) -> str:
        return (datos.get(k) or "").strip()

    empresa   = g("empresa")
    sucursal  = g("sucursal")
    cc        = g("centro_costos")
    direccion = g("direccion")
    proveedor = g("proveedor_nombre")
    motivo    = g("motivo_pago")
    folio     = g("folio_cfdi")
    nota_cred = g("notas_credito")
    try:
        monto = float(str(datos.get("monto_total") or 0).replace(",", ""))
    except Exception:
        monto = 0.0
    monto_str = f"$ {monto:,.2f}" if monto else ""
    letra     = g("importe_letra")
    banco     = g("banco")
    clabe     = g("clabe")
    no_cta    = g("no_cuenta")
    obs       = g("observaciones")
    mes_pres  = g("mes_presupuesto")
    mes_pago_ = g("mes_pago")
    fec_sol   = g("fecha_proceso") or date.today().strftime("%d/%m/%Y")
    analista  = g("analista_nombre")
    gerente   = g("gerente_nombre")
    visto_bno = g("visto_bno")
    depto_fin = g("depto_finanzas")
    dir_fin   = g("dir_financiera")
    dir_gral  = g("dir_general")

    logo_path = _get_logo(empresa)

    buf = io.BytesIO()
    cv  = canvas.Canvas(buf, pagesize=letter)

    # ── Fondo (plantilla Excel → JPG 1275×1650 px) ──────────────────
    plantilla = _plantilla()
    if plantilla:
        cv.drawImage(plantilla, 0, 0, width=PW, height=PH, preserveAspectRatio=False)

    # ── Logo empresa (área B1:E6 del Excel, top-left) ────────────────
    if logo_path:
        try:
            cv.drawImage(logo_path, 22, PH - 88, width=95, height=52,
                         preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    elif empresa:
        cv.setFont("Helvetica", 7.5)
        cv.setFillColor(NEGRO)
        cv.drawString(22, PH - 62, empresa)

    # ── Helpers ──────────────────────────────────────────────────────
    def ds(txt, x, y, sz=8.5, bold=False):
        if not txt:
            return
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.setFillColor(NEGRO)
        cv.drawString(x, y, str(txt))

    def dsr(txt, x, y, sz=8.5, bold=False):
        if not txt:
            return
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.setFillColor(NEGRO)
        cv.drawRightString(x, y, str(txt))

    def dsc(txt, x, y, sz=7.0, bold=False):
        if not txt:
            return
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.setFillColor(NEGRO)
        cv.drawCentredString(x, y, str(txt))

    # ── DATOS PRINCIPALES ────────────────────────────────────────────
    # RL y = 792 - y_top_excel - ascender(sz)
    # ascender Helvetica: 8.5pt→9, 8pt→9, 7.5pt→8, 11pt→12, 6.5pt→7

    # Sucursal / Fecha de solicitud  (row 7, y_excel≈105 → RL y=678)
    ds(sucursal, 143.0, 678.0, sz=8.5, bold=True)
    ds(fec_sol,  481.0, 679.0, sz=8.5)

    # Centro de Costos / Dirección  (row 9, y_excel≈120 → RL y=663)
    ds(cc,        145.0, 663.0, sz=8.0)
    ds(direccion, 464.0, 663.0, sz=8.0, bold=True)

    # Beneficiario  (row 12, y_excel≈140 → RL y=643)
    ds(proveedor, 166.0, 643.0, sz=8.5, bold=True)

    # Motivo de pago  (row 14, y_excel≈153 → RL y=630)
    ds(motivo, 166.0, 630.0, sz=8.5)

    # Folio CFDI / Nota de crédito  (row 19, y_excel≈194 → RL y=589)
    ds(folio,     240.0, 589.0, sz=8.5, bold=True)
    ds(nota_cred, 440.0, 589.0, sz=8.0)

    # Importe en letra / Monto  (row 25, y_excel≈241; sz distintas → RL y distintos)
    ds(letra,      236.0, 543.0, sz=7.5)           # 792-241-8=543
    dsr(monto_str, 562.0, 539.0, sz=11.0, bold=True)  # 792-241-12=539

    # Banco / CLABE / No. de Cuenta  (row 30, y_excel≈275 → RL y=508)
    ds(banco,  110.0, 508.0, sz=8.5, bold=True)
    ds(clabe,  270.0, 508.0, sz=8.5, bold=True)
    ds(no_cta, 466.0, 508.0, sz=8.5, bold=True)

    # Observaciones  (row 32, y_excel≈298 → RL y=485)
    ds(obs, 166.0, 485.0, sz=8.0, bold=True)

    # ── EXCLUSIVO FINANZAS ───────────────────────────────────────────
    # Mes presupuesto / Mes pago  (row 42, y_excel≈378 → RL y=405)
    ds(mes_pres,  296.0, 405.0, sz=8.5, bold=True)
    ds(mes_pago_, 477.0, 405.0, sz=8.5, bold=True)

    # CC Finanzas  (row 45, y_excel≈403 → RL y=380)
    ds(cc, 278.0, 380.0, sz=8.0)

    # ── FIRMAS ───────────────────────────────────────────────────────
    # Roles/etiquetas vienen del background (Excel); aquí solo los nombres.
    XC1, XC2, XC3, XC4 = 117.0, 220.0, 403.0, 505.0

    # Fila 1 — Analista / Gerente  (row 63, y_excel≈535 → RL y=250)
    dsc(analista, XC1, 250.0, sz=6.0, bold=True)
    dsc(gerente,  XC2, 250.0, sz=6.0, bold=True)

    # Fila 2 — Vo.Bo. / Depto.Fin / Dir.Fin / Dir.Gral  (row 76, y_excel≈632 → RL y=153)
    dsc(visto_bno, XC1, 153.0, sz=6.0, bold=True)
    dsc(depto_fin, XC2, 153.0, sz=6.0, bold=True)
    dsc(dir_fin,   XC3, 153.0, sz=6.0, bold=True)
    dsc(dir_gral,  XC4, 153.0, sz=6.0, bold=True)

    cv.save()
    return buf.getvalue()


def exportar_estado_cuenta_xlsx(estado: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    neg = PatternFill("solid", fgColor="374151")
    ver = PatternFill("solid", fgColor="166534")
    amb = PatternFill("solid", fgColor="92400E")
    grs = PatternFill("solid", fgColor="F3F4F6")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Estado de Cuenta — {estado.get('mes_nombre', '')} {estado.get('anio', '')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = neg
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Total pagado: ${estado.get('total_pagado', 0):,.2f}"
    ws["A2"].font = Font(bold=True, color="15803D")
    ws.merge_cells("E2:G2")
    ws["E2"] = f"Total pendiente: ${estado.get('total_pendiente', 0):,.2f}"
    ws["E2"].font = Font(bold=True, color="B45309")

    hdrs = ["ID", "Proveedor", "Empresa", "Motivo de pago", "Monto", "Estatus", "Fecha"]
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = neg
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for ri, pg in enumerate(estado.get("todos", []), 4):
        est = (pg.get("estatus") or "PENDIENTE").upper()
        ef = ver if est == "PAGADO" else amb
        for col, v in enumerate([
            pg.get("id", ""), pg.get("proveedor_nombre", ""),
            pg.get("empresa", ""), pg.get("motivo_pago", ""),
            pg.get("monto_total", 0), est, pg.get("fecha_proceso", ""),
        ], 1):
            cell = ws.cell(row=ri, column=col, value=v)
            cell.fill = ef if col == 6 else (grs if ri % 2 == 0 else PatternFill())
            cell.border = thin
            if col == 5:
                cell.number_format = "$#,##0.00"

    for col, w in enumerate([8, 25, 22, 35, 14, 12, 14], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
