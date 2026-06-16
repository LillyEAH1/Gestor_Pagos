"""
exportar.py v60
Plantilla de fondo + drawString con coordenadas medidas pixel a pixel
del ejemplo Solicitud_B&B_TELEFONOS_DE_MEXICO_140526_MAY.
Imagen 1275×1650 px → Letter 612×792 pts, ratio 0.48, Y invertido.
CERO defaults de empresa. Campo vacío = no imprime nada.
"""
from __future__ import annotations
import json
from pathlib import Path
from datetime import date

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

PW, PH = letter          # 612 × 792 pts
NEGRO  = colors.black
MUTED  = colors.HexColor("#333333")


def _logos_dir() -> Path:
    return Path(__file__).parent / "logos"


def _plantilla() -> str:
    p = _logos_dir() / "formato2_page-0001.jpg"
    return str(p) if p.exists() else ""


def _get_logo(empresa: str):
    if not empresa:
        return None
    mf = _logos_dir() / "logo_map.json"
    if not mf.exists():
        return None
    try:
        m = json.loads(mf.read_text(encoding="utf-8"))
        eu = empresa.upper()
        for k, v in m.items():
            if k.upper() in eu or eu[:12] in k.upper():
                p = _logos_dir() / v
                if p.exists():
                    return str(p)
    except Exception:
        pass
    return None


def exportar_solicitud_pdf(datos: dict, ruta_salida: str) -> None:
    """
    Genera el PDF con la plantilla como fondo y los datos en
    coordenadas fijas. NUNCA pone empresa por default.
    """

    def g(k: str) -> str:
        return (datos.get(k) or "").strip()

    # ── Extraer datos — CERO defaults ─────────────────────────────────────
    empresa   = g("empresa")
    sucursal  = g("sucursal")
    cc        = g("centro_costos")
    direccion = g("direccion")
    proveedor = g("proveedor_nombre")
    motivo    = g("motivo_pago")
    folio     = g("folio_cfdi")
    nota_cred = g("notas_credito")
    try:
        monto = float(str(datos.get("monto_total") or 0).replace(",", ""))
    except Exception:
        monto = 0.0
    monto_str = f"$ {monto:,.2f}" if monto else ""
    letra     = g("importe_letra")
    banco     = g("banco")
    clabe     = g("clabe")
    no_cta    = g("no_cuenta")
    obs       = g("observaciones")
    mes_pres  = g("mes_presupuesto")
    mes_pago_ = g("mes_pago")
    fec_sol   = g("fecha_proceso") or date.today().strftime("%d/%m/%Y")
    analista  = g("analista_nombre")
    gerente   = g("gerente_nombre")
    visto_bno = g("visto_bno")
    depto_fin = g("depto_finanzas")
    dir_fin   = g("dir_financiera")
    dir_gral  = g("dir_general")

    logo_path = _get_logo(empresa)

    # ── Canvas ────────────────────────────────────────────────────────────
    cv = canvas.Canvas(ruta_salida, pagesize=letter)

    # ══ PASO 1: Plantilla como fondo completo ═════════════════════════════
    plantilla = _plantilla()
    if plantilla:
        cv.drawImage(plantilla, 0, 0, width=PW, height=PH,
                     preserveAspectRatio=False)

    # ══ PASO 2: Logo empresa (zona superior izquierda del formato) ════════
    if logo_path:
        try:
            cv.drawImage(logo_path, 22, PH - 88, width=95, height=52,
                         preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    elif empresa:
        cv.setFont("Helvetica", 7.5)
        cv.setFillColor(NEGRO)
        cv.drawString(22, PH - 62, empresa)

    # ══ PASO 3: Texto sobre plantilla — coordenadas fijas ═════════════════
    # Derivadas de imagen 1275×1650 px → pdf_x = px*0.48, pdf_y = 792-py*0.48
    # Verificadas campo a campo contra el ejemplo lleno.

    def ds(txt: str, x: float, y: float,
           sz: float = 8.5, bold: bool = False,
           color=NEGRO) -> None:
        """drawString — nunca imprime vacío ni None."""
        if not txt:
            return
        cv.setFillColor(color)
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.drawString(x, y, str(txt))
        cv.setFillColor(NEGRO)

    def dsr(txt: str, x: float, y: float,
            sz: float = 8.5, bold: bool = False) -> None:
        """drawRightString alineado a la derecha."""
        if not txt:
            return
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.setFillColor(NEGRO)
        cv.drawRightString(x, y, str(txt))

    def dsc(txt: str, x: float, y: float,
            sz: float = 7.0, bold: bool = False,
            color=NEGRO) -> None:
        """drawCentredString."""
        if not txt:
            return
        cv.setFillColor(color)
        cv.setFont("Helvetica-Bold" if bold else "Helvetica", sz)
        cv.drawCentredString(x, y, str(txt))
        cv.setFillColor(NEGRO)

    # ── SUCURSAL | Fecha de solicitud ─────────────────────────────────────
    # px=(376,248) → pdf=(180.5, 673.0)
    ds(sucursal, 180.5, 673.0, sz=8.5, bold=True)
    # px=(1030,248) → pdf=(494.4, 673.0)
    ds(fec_sol,  494.4, 673.0, sz=8.5)

    # ── CENTRO DE COSTOS | DIRECCIÓN ──────────────────────────────────────
    # px=(287,300) → pdf=(137.8, 648.0)
    ds(cc,        137.8, 648.0, sz=8.0)
    # px=(950,300) → pdf=(456.0, 648.0)
    ds(direccion, 456.0, 648.0, sz=8.0, bold=True)

    # ── BENEFICIARIO ──────────────────────────────────────────────────────
    # px=(287,373) → pdf=(137.8, 613.0)
    ds(proveedor, 137.8, 613.0, sz=8.5, bold=True)

    # ── MOTIVO DE PAGO ────────────────────────────────────────────────────
    # px=(287,418) → pdf=(137.8, 591.4)
    ds(motivo, 137.8, 591.4, sz=8.5)

    # ── DATOS DE CFDI — Folio ─────────────────────────────────────────────
    # El folio va en la celda central "CFDI". Centrado en x≈340 (entre 221 y 459)
    # px=(560,548) → pdf=(268.8, 529.0) — usar centro de la celda CFDI
    dsc(folio,     340.0, 529.0, sz=8.5, bold=True)
    # Nota de crédito en la columna derecha px=(800,548) → pdf=(384,529)
    ds(nota_cred,  468.0, 529.0, sz=8.0)

    # ── DATOS DE PAGO — Importe ───────────────────────────────────────────
    # Importe en letra px=(287,688) → pdf=(137.8, 461.8)
    ds(letra, 137.8, 461.8, sz=7.5)
    # Monto alineado a derecha del recuadro (borde derecho ≈ 598 pts)
    # px=(1245,688) → pdf=(597.6, 461.8)
    dsr(monto_str, 597.6, 461.8, sz=11.0, bold=True)

    # ── Banco | CLABE | No. de Cuenta ─────────────────────────────────────
    # px=(170,742) → pdf=(81.6, 435.8)
    ds(banco,  81.6,  435.8, sz=8.5, bold=True)
    # px=(490,742) → pdf=(235.2, 435.8)
    ds(clabe,  235.2, 435.8, sz=8.5, bold=True)
    # px=(1050,742) → pdf=(504.0, 435.8)
    ds(no_cta, 504.0, 435.8, sz=8.5, bold=True)

    # ── Observaciones ─────────────────────────────────────────────────────
    # px=(287,798) → pdf=(137.8, 409.0)
    ds(obs, 137.8, 409.0, sz=8.0, bold=True)

    # ── EXCLUSIVO FINANZAS ────────────────────────────────────────────────
    # Mes del Presupuesto px=(455,958) → pdf=(218.4, 332.2)
    ds(mes_pres,  218.4, 332.2, sz=8.5, bold=True)
    # Mes del Pago px=(895,958) → pdf=(429.6, 332.2)
    ds(mes_pago_, 429.6, 332.2, sz=8.5, bold=True)
    # CC Finanzas px=(287,1030) → pdf=(137.8, 297.6)
    ds(cc, 137.8, 297.6, sz=8.0)

    # ══ FIRMAS ════════════════════════════════════════════════════════════
    # Centros X de las 4 columnas de firma, derivados de la imagen:
    # Col 1 centro: px≈158 → pdf=75.8 — ancho col ≈ 153 pts → centro ≈ 76+76=152? 
    # Usar los centros reales medidos:
    # C1=158px→76, C2=476px→228, C3=795px→382, C4=1115px→535
    # Ajuste: usar centro de cada cuarto de página
    # Cuartos: [0-153], [153-306], [306-459], [459-612] → centros: 76, 229, 382, 535
    XC1, XC2, XC3, XC4 = 76.5, 229.5, 382.5, 535.5

    # ── Fila 1 — Nombres SOBRE la línea, Depts BAJO la línea ─────────────
    # Línea de firma a px_y≈1157 → pdf_y=237.0
    # Nombre sobre línea: pdf_y = 237 + 5 = 242 → 233.8 (medido del ejemplo)
    # Dept bajo línea: pdf_y = 237 - 11 = 226 → 226.6 (medido)
    Y1N = 233.8   # y nombre fila 1
    Y1D = 226.6   # y dept fila 1

    dsc(analista,              XC1, Y1N, sz=6.5, bold=True)
    dsc(gerente,               XC2, Y1N, sz=6.5, bold=True)
    # Cols 3 y 4 no tienen nombre configurado — quedan en blanco (ya tiene NOMBRE en plantilla)

    dsc("ANALISTA DE SISTEMAS", XC1, Y1D, sz=5.5, color=MUTED)
    dsc("GERENTE DE SISTEMAS",  XC2, Y1D, sz=5.5, color=MUTED)
    # Cols 3 y 4 ya tienen "DEPTO. / NOMBRE" en la plantilla — no sobreescribir

    # ── Fila 2 — Nombres SOBRE la línea, Depts BAJO la línea ─────────────
    # Línea de firma 2 a px_y≈1383 → pdf_y=128.2
    # Nombre: 128.2 + 5 = 133 → 125.8 (medido)
    # Dept:   128.2 - 10 = 118 → 118.6 (medido)
    Y2N = 125.8   # y nombre fila 2
    Y2D = 118.6   # y dept fila 2

    dsc(visto_bno, XC1, Y2N, sz=6.5, bold=True)
    dsc(depto_fin, XC2, Y2N, sz=6.5, bold=True)
    dsc(dir_fin,   XC3, Y2N, sz=6.5, bold=True)
    dsc(dir_gral,  XC4, Y2N, sz=6.5, bold=True)

    dsc("DEPARTAMENTO DE FINANZAS", XC1, Y2D, sz=5.5, color=MUTED)
    dsc("DEPARTAMENTO DE FINANZAS", XC2, Y2D, sz=5.5, color=MUTED)
    dsc("DIRECCIÓN FINANCIERA",     XC3, Y2D, sz=5.5, color=MUTED)
    dsc("DIRECCIÓN GENERAL",        XC4, Y2D, sz=5.5, color=MUTED)

    cv.save()


def exportar_estado_cuenta_xlsx(estado: dict, ruta_salida: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    neg  = PatternFill("solid", fgColor="374151")
    ver  = PatternFill("solid", fgColor="166534")
    amb  = PatternFill("solid", fgColor="92400E")
    grs  = PatternFill("solid", fgColor="F3F4F6")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Estado de Cuenta — {estado.get('mes_nombre','')} {estado.get('anio','')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = neg
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Total pagado: ${estado.get('total_pagado', 0):,.2f}"
    ws["A2"].font = Font(bold=True, color="15803D")
    ws.merge_cells("E2:G2")
    ws["E2"] = f"Total pendiente: ${estado.get('total_pendiente', 0):,.2f}"
    ws["E2"].font = Font(bold=True, color="B45309")

    hdrs = ["ID", "Proveedor", "Empresa", "Motivo de pago", "Monto", "Estatus", "Fecha"]
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = neg
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for ri, pg in enumerate(estado.get("todos", []), 4):
        est = (pg.get("estatus") or "PENDIENTE").upper()
        ef  = ver if est == "PAGADO" else amb
        for col, v in enumerate([
            pg.get("id", ""), pg.get("proveedor_nombre", ""),
            pg.get("empresa", ""), pg.get("motivo_pago", ""),
            pg.get("monto_total", 0), est, pg.get("fecha_proceso", "")
        ], 1):
            cell = ws.cell(row=ri, column=col, value=v)
            cell.fill = ef if col == 6 else (grs if ri % 2 == 0 else PatternFill())
            cell.border = thin
            if col == 5:
                cell.number_format = "$#,##0.00"

    for col, w in enumerate([8, 25, 22, 35, 14, 12, 14], 1):
        ws.column_dimensions[chr(64 + col)].width = w
    wb.save(ruta_salida)
